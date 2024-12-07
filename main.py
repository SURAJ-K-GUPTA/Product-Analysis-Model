from fastapi import FastAPI, File, UploadFile, HTTPException
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from PIL import Image
import openpyxl
import re
import os
import torch
import gc
from io import BytesIO
from transformers import Qwen2VLForConditionalGeneration, AutoProcessor
import multiprocessing
import signal

# Cleanup semaphore resources (to prevent leaked semaphore warnings)
def cleanup_semaphores():
    try:
        multiprocessing.resource_tracker._semaphore_tracker._cache.clear()
    except AttributeError:
        pass
    signal.signal(signal.SIGTERM, signal.SIG_DFL)

cleanup_semaphores()

app = FastAPI()

# Mount the frontend directory as static files
app.mount("/static", StaticFiles(directory="frontend"), name="static")

# Load the model and processor during startup
@app.on_event("startup")
def load_model_and_initialize():
    global model, processor
    try:
        model = Qwen2VLForConditionalGeneration.from_pretrained(
            "Qwen/Qwen2-VL-2B-Instruct",
            torch_dtype="auto",
            device_map="auto",
            low_cpu_mem_usage=True,
        )
        processor = AutoProcessor.from_pretrained("Qwen/Qwen2-VL-2B-Instruct")
    except Exception as e:
        print(f"Error loading model: {e}")
        raise RuntimeError("Model loading failed. Check your configuration.")

    # Ensure Excel file exists
    file_path = "product_analysis.xlsx"
    if not os.path.exists(file_path):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Product Analysis"
        headers = ["Product Name", "Category", "Quantity", "Count", "Expiry Date", "Freshness Index", "Shelf Life"]
        sheet.append(headers)
        workbook.save(file_path)

# For packaged products
packaged_product_pattern = r"\*\*Product Name:\*\* (.*?)\n\s*- \*\*Product Category:\*\* (.*?)\n\s*- \*\*Product Quantity:\*\* (.*?)\n\s*- \*\*Product Count:\*\* (.*?)\n\s*- \*\*Expiry Date:\*\* (.*?)\n"

# For fruits and vegetables
fruits_vegetables_pattern = r"\*\*Type of fruit/vegetable:\*\* (.*?)\n\s*- \*\*Freshness Index:\*\* (.*?)\n\s*- \*\*Estimated Shelf Life:\*\* (.*?)\n"

@app.get("/", response_class=HTMLResponse)
async def serve_homepage():
    try:
        with open("frontend/index.html", "r") as f:
            return f.read()
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error loading homepage: {e}")

@app.post("/analyze-image/")
async def analyze_image(file: UploadFile = File(...)):
    print("Image analysis route triggered")
    try:
        # Log file metadata
        print(f"File name: {file.filename}")
        print(f"Content type: {file.content_type}")

        # Validate file type
        if file.content_type not in ["image/jpeg", "image/png"]:
            raise HTTPException(status_code=400, detail="Invalid file type. Please upload a JPEG or PNG image.")

        # Read file contents
        contents = await file.read()
        print(f"File content (first 100 bytes): {contents[:100]}")

        # Process the image with Pillow
        try:
            # Verify and reload the image
            image = Image.open(BytesIO(contents))
            image.verify()
            image = Image.open(BytesIO(contents))
            image = image.resize((512, 512))
        except Exception as e:
            print(f"Image processing error: {e}")
            raise HTTPException(status_code=400, detail=f"Invalid image file: {e}")

        # Prepare input for the model
        messages = [
            {
                "role": "user",
                "content": [
                    {"type": "image", "image_url": "Captured from webcam"},
                    {"type": "text",
                     "text": """This image contains fruits, vegetables, or packaged products.
                        Please analyze the image and provide:
                        - For packaged products:
                            - Product Name
                            - Product Category
                            - Product Quantity
                            - Product Count
                            - Expiry Date (if available)
                        - For fruits and vegetables:
                            - Type of fruit/vegetable
                            - Freshness Index (based on visual cues)
                            - Estimated Shelf Life"""
                     }
                ]
            }
        ]
        text_prompt = processor.apply_chat_template(messages, add_generation_prompt=True)
        inputs = processor(
            text=[text_prompt],
            images=[image],
            padding=True,
            return_tensors="pt"
        )

        # Ensure inputs are on the same device as the model
        model_device = next(model.parameters()).device
        inputs = inputs.to(model_device)

        # Generate output using the model
        with torch.no_grad():
            output_ids = model.generate(**inputs, max_new_tokens=1024)
        output_text = processor.batch_decode(
            output_ids, skip_special_tokens=True, clean_up_tokenization_spaces=True
        )[0]
        print(f"Model Output: {output_text}")  # Debug the model's output

        # Parse output text to extract relevant fields
        # Parse output text to extract relevant fields
        product_name = category = quantity = count = expiry_date = freshness_index = shelf_life = "N/A"
        try:
            # Match packaged product details
            packaged_product_match = re.search(packaged_product_pattern, output_text)
            if packaged_product_match:
                product_name = packaged_product_match.group(1).strip()
                category = packaged_product_match.group(2).strip()
                quantity = packaged_product_match.group(3).strip()
                count = packaged_product_match.group(4).strip()
                expiry_date = packaged_product_match.group(5).strip()

            # Match fruits and vegetables details
            fruits_vegetables_match = re.search(fruits_vegetables_pattern, output_text)
            if fruits_vegetables_match:
                product_name = fruits_vegetables_match.group(1).strip()
                category = "Fruit/Vegetable"
                freshness_index = fruits_vegetables_match.group(2).strip()
                shelf_life = fruits_vegetables_match.group(3).strip()
        except Exception as e:
            print(f"Error parsing output: {e}")

        # If no data was matched, log a warning
        if product_name == "N/A" and freshness_index == "N/A":
            print("Warning: No data matched the expected patterns.")

        # Update Excel with parsed results
        file_path = "product_analysis.xlsx"
        try:
            workbook = openpyxl.load_workbook(file_path)
        except Exception as e:
            print("Workbook load failed, recreating product_analysis.xlsx")
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Product Analysis"
            headers = ["Product Name", "Category", "Quantity", "Count", "Expiry Date", "Freshness Index", "Shelf Life"]
            sheet.append(headers)
            workbook.save(file_path)
            workbook = openpyxl.load_workbook(file_path)

        sheet = workbook.active
        sheet.append([product_name, category, quantity, count, expiry_date, freshness_index, shelf_life])
        workbook.save(file_path)

        # Return parsed data as API response
        return {
            "message": "Analysis completed successfully",
            "data": {
                "Product Name": product_name,
                "Category": category,
                "Quantity": quantity,
                "Count": count,
                "Expiry Date": expiry_date,
                "Freshness Index": freshness_index,
                "Shelf Life": shelf_life
            }
        }

    except Exception as e:
        print(f"Backend error: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})
@app.get("/download-excel/")
async def download_excel():
    file_path = "product_analysis.xlsx"
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="Excel file not found.")
    return FileResponse(
        file_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="product_analysis.xlsx",
    )
